import requests
import json
from openpyxl import load_workbook
from saveContent import *
from savePackageData import *
from getPackageTypeApi import *
from getMediaApi import *
from offerApi import *
from getContent import *
from getVodFilters import *

wb = load_workbook(r'C:\Users\MTMUNIAN\Desktop\Python\Excel\AMS.xlsx')

sheet = wb.active

i = 3
for row in sheet.iter_rows(min_row=4, max_col=1):
    for cell in row:
        i += 1

        # #################--- IF status = MP ---##################
        if sheet['A' + str(i)].value == 'MP':

            # #################--- (MP) Title ---##################
            if sheet['J' + str(i)].value is None and sheet['K' + str(i)].value is None:

                # #################################--- contentSave API Call Function---##################################

                # #################-- Set contentId 0 for new content or update if ACID exist---##################
                if sheet['C' + str(i)].value is None:
                    contentid = 0
                else:
                    contentid = sheet['C' + str(i)].value

                systemtitle = sheet['E' + str(i)].value

                # #################-- Set Episodic value 0 or 1---##################
                if sheet['F' + str(i)].value == 'Y':
                    episodic = 1
                else:
                    episodic = 0

                if not sheet['G' + str(i)].value is None:
                    recordtype = sheet['G' + str(i)].value
                else:
                    recordtype = ''

                if not sheet['H' + str(i)].value is None:
                    yearofrelease = sheet['H' + str(i)].value
                else:
                    yearofrelease = ''

                if not sheet['I' + str(i)].value is None:
                    originallanguage = sheet['I' + str(i)].value
                else:
                    originallanguage = ''

                # #################-- Set Synopsis---##################
                contentsynopsis = []
                if not sheet['L' + str(i)].value is None:
                    shtSynopEng = {"contentSynopsisId": "0", "synopsisType": "Short", "synopsis": "" + sheet['L' + str(i)].value + "", "synopsisLanguage": "English"}
                    contentsynopsis = [shtSynopEng]
                if not sheet['M' + str(i)].value is None:
                    lngSynopEng = {"contentSynopsisId": "0", "synopsisType": "Long", "synopsis": "" + sheet['M' + str(i)].value + "", "synopsisLanguage": "English"}
                    contentsynopsis = [shtSynopEng, lngSynopEng]
                if not sheet['N' + str(i)].value is None:
                    shtSynopMly = {"contentSynopsisId": "0", "synopsisType": "Short", "synopsis": "" + sheet['N' + str(i)].value + "", "synopsisLanguage": "Malay"}
                    contentsynopsis = [shtSynopEng, lngSynopEng, shtSynopMly]
                if not sheet['O' + str(i)].value is None:
                    lngSynopMly = {"contentSynopsisId": "0", "synopsisType": "Long", "synopsis": "" + sheet['O' + str(i)].value + "", "synopsisLanguage": "Malay"}
                    contentsynopsis = [shtSynopEng, lngSynopEng, shtSynopMly, lngSynopMly]
                if not sheet['P' + str(i)].value is None:
                    shtSynopChi = {"contentSynopsisId": "0", "synopsisType": "Short", "synopsis": "" + sheet['P' + str(i)].value + "", "synopsisLanguage": "Chinese"}
                    contentsynopsis = [shtSynopEng, lngSynopEng, shtSynopMly, lngSynopMly, shtSynopChi]
                if not sheet['Q' + str(i)].value is None:
                    lngSynopChi = {"contentSynopsisId": "0", "synopsisType": "Long", "synopsis": "" + sheet['Q' + str(i)].value + "", "synopsisLanguage": "Chinese"}
                    contentsynopsis = [shtSynopEng, lngSynopEng, shtSynopMly, lngSynopMly, shtSynopChi, lngSynopChi]
                if not sheet['R' + str(i)].value is None:
                    shtSynopTam = {"contentSynopsisId": "0", "synopsisType": "Short", "synopsis": "" + sheet['R' + str(i)].value + "", "synopsisLanguage": "Tamil"}
                    contentsynopsis = [shtSynopEng, lngSynopEng, shtSynopMly, lngSynopMly, shtSynopChi, lngSynopChi, shtSynopTam]
                if not sheet['S' + str(i)].value is None:
                    lngSynopTam = {"contentSynopsisId": "0", "synopsisType": "Long", "synopsis": "" + sheet['S' + str(i)].value + "", "synopsisLanguage": "Tamil"}
                    contentsynopsis = [shtSynopEng, lngSynopEng, shtSynopMly, lngSynopMly, shtSynopChi, lngSynopChi, shtSynopTam, lngSynopTam]
                if not sheet['T' + str(i)].value is None:
                    shtSynopInd = {"contentSynopsisId": "0", "synopsisType": "Short", "synopsis": "" + sheet['T' + str(i)].value + "", "synopsisLanguage": "Bahasa Indonesia"}
                    contentsynopsis = [shtSynopEng, lngSynopEng, shtSynopMly, lngSynopMly, shtSynopChi, lngSynopChi, shtSynopTam, lngSynopTam, shtSynopInd]
                if not sheet['U' + str(i)].value is None:
                    lngSynopInd = {"contentSynopsisId": "0", "synopsisType": "Long", "synopsis": "" + sheet['U' + str(i)].value + "", "synopsisLanguage": "Bahasa Indonesia"}
                    contentsynopsis = [shtSynopEng, lngSynopEng, shtSynopMly, lngSynopMly, shtSynopChi, lngSynopChi, shtSynopTam, lngSynopTam, shtSynopInd, lngSynopInd]
                if not sheet['V' + str(i)].value is None:
                    shtSynopTh = {"contentSynopsisId": "0", "synopsisType": "Short", "synopsis": "" + sheet['V' + str(i)].value + "", "synopsisLanguage": "Thai"}
                    contentsynopsis = [shtSynopEng, lngSynopEng, shtSynopMly, lngSynopMly, shtSynopChi, lngSynopChi, shtSynopTam, lngSynopTam, shtSynopInd, lngSynopInd, shtSynopTh]
                if not sheet['W' + str(i)].value is None:
                    lngSynopTh = {"contentSynopsisId": "0", "synopsisType": "Long", "synopsis": "" + sheet['W' + str(i)].value + "", "synopsisLanguage": "Thai"}
                    contentsynopsis = [shtSynopEng, lngSynopEng, shtSynopMly, lngSynopMly, shtSynopChi, lngSynopChi, shtSynopTam, lngSynopTam, shtSynopInd, lngSynopInd, shtSynopTh, lngSynopTh]

                # #################-- Content Title---##################
                contenttitle = []
                if not sheet['X' + str(i)].value is None:
                    titleEng = {"contentTitleId": "0", "title": "" + sheet['X' + str(i)].value + "", "titleType": "Standard", "titleLanguage": "English"}
                    contenttitle = [titleEng]
                if not sheet['Y' + str(i)].value is None:
                    titleMly = {"contentTitleId": "0", "title": "" + sheet['Y' + str(i)].value + "", "titleType": "Standard", "titleLanguage": "Malay"}
                    contenttitle = [titleEng, titleMly]
                if not sheet['Z' + str(i)].value is None:
                    titleChi = {"contentTitleId": "0", "title": "" + sheet['Z' + str(i)].value + "", "titleType": "Standard", "titleLanguage": "Chinese"}
                    contenttitle = [titleEng, titleMly, titleChi]
                if not sheet['AA' + str(i)].value is None:
                    titleTam = {"contentTitleId": "0", "title": "" + sheet['AA' + str(i)].value + "", "titleType": "Standard", "titleLanguage": "Tamil"}
                    contenttitle = [titleEng, titleMly, titleChi, titleTam]
                if not sheet['AB' + str(i)].value is None:
                    titleInd = {"contentTitleId": "0", "title": "" + sheet['AB' + str(i)].value + "", "titleType": "Standard", "titleLanguage": "Bahasa Indonesia"}
                    contenttitle = [titleEng, titleMly, titleChi, titleTam, titleInd]
                if not sheet['AC' + str(i)].value is None:
                    titleTh = {"contentTitleId": "0", "title": "" + sheet['AC' + str(i)].value + "", "titleType": "Standard", "titleLanguage": "Thai"}
                    contenttitle = [titleEng, titleMly, titleChi, titleTam, titleInd, titleTh]

                # #################-- Content Genre---##################
                if not sheet['AD' + str(i)].value is None and not sheet['AE' + str(i)].value is None:
                    genreDVB = {"contentGenreId": "0", "genreType": "DVB", "genre": "" + sheet['AD' + str(i)].value + "", "subgenre": "" + sheet['AE' + str(i)].value}
                    contentgenre = [genreDVB]
                else:
                    print('Genre is missing')
                    break

                if not sheet['AF' + str(i)].value is None and not sheet['AG' + str(i)].value is None:
                    filter1 = {"contentGenreId": "0", "genreType": "Category/Filter", "genre": "" + sheet['AF' + str(i)].value + "", "subgenre": "" + sheet['AG' + str(i)].value}
                    contentgenre = [genreDVB, filter1]
                else:
                    print('Filter1 is missing')
                    break

                if not sheet['AH' + str(i)].value is None and not sheet['AI' + str(i)].value is None:
                    filter2 = {"contentGenreId": "0", "genreType": "Category/Filter", "genre": "" + sheet['AH' + str(i)].value + "", "subgenre": "" + sheet['AI' + str(i)].value}
                    contentgenre = [genreDVB, filter1, filter2]
                else:
                    print('Filter2 is missing')
                    break

                # #################-- Content Talent---##################
                contenttalent = []
                if not sheet['AJ' + str(i)].value is None:
                    talActEng = {"contentTalentId": "0", "talentType": "Actor", "talentName": "" + sheet['AJ' + str(i)].value + "", "talentLanguage": "English", "billingOrder": "0"}
                    contenttalent = [talActEng]
                if not sheet['AK' + str(i)].value is None:
                    talActMly = {"contentTalentId": "0", "talentType": "Actor", "talentName": "" + sheet['AK' + str(i)].value + "", "talentLanguage": "Malay", "billingOrder": "0"}
                    contenttalent = [talActEng, talActMly]
                if not sheet['AL' + str(i)].value is None:
                    talActChi = {"contentTalentId": "0", "talentType": "Actor", "talentName": "" + sheet['AL' + str(i)].value + "", "talentLanguage": "Chinese", "billingOrder": "0"}
                    contenttalent = [talActEng, talActMly, talActChi]
                if not sheet['AM' + str(i)].value is None:
                    talActTam = {"contentTalentId": "0", "talentType": "Actor", "talentName": "" + sheet['AM' + str(i)].value + "", "talentLanguage": "Tamil", "billingOrder": "0"}
                    contenttalent = [talActEng, talActMly, talActChi, talActTam]

                # ########## saveContentapi API call function ###############
                contentRes = saveContentapi(contentid, systemtitle, episodic, recordtype, yearofrelease, originallanguage, contentsynopsis, contenttitle, contentgenre, contenttalent)
                print(contentRes['status'], contentRes['responseMessage'])
                contentid = contentRes['status']

                # ################### ----- Save Excel with changes ACID Number------- ####################
                sheet['C' + str(i)] = contentid

                # #################-- Set pkgId for new package or update if pkgId exist---##################
                if sheet['AR' + str(i)].value is None:
                    pkgid = 0
                else:
                    pkgid = sheet['AR' + str(i)].value

                # ################### ----- Prepare package name PackageType + PackageName + ChannelOwner------- ####################
                if not sheet['AS' + str(i)].value is None and not sheet['B' + str(i)].value is None and not sheet['AU' + str(i)].value is None:
                    packagename = sheet['AS' + str(i)].value + ' - ' + sheet['B' + str(i)].value + ' (' + sheet['AU' + str(i)].value + ')'
                    # print(packagename)
                else:
                    print("Package Type | Package Name | ChannelOwner is Empty")
                    break

                # ################### ----- Set packagetype------- ####################
                packageTypeRes = getPackageTypeApi()
                for packageType in packageTypeRes['packageType']:
                    if packageType['packageName'] == sheet['AS' + str(i)].value:
                        pkgtypeid = packageType['packageId']
                        # print(packageType['packageId'])

                # ################### ----- Set channelOwner  (single and multiple value)------- ####################
                if not sheet['AU' + str(i)].value is None:
                    channelOwnerExcel = sheet['AU' + str(i)].value
                    channelOwnerExcelDict = channelOwnerExcel.split("|")
                    channelownerArray = []
                    for channelOwn in channelOwnerExcelDict:
                        # print(channelOwn)
                        channelOwnerRes = getPackageTypeApi()
                        for channelOwner in channelOwnerRes['channels']:
                            if channelOwner['channelName'] == channelOwn:
                                # print(channelOwner['channelId'])
                                channelownerArray.append("" + str(channelOwner['channelId']) + "")
                    channelowner = channelownerArray
                    # print(channelowner)
                else:
                    print('Channel Owner is Empty')
                    break

                # ################### ----- Set autoPublish------ ####################
                if sheet['AT' + str(i)].value == 'Y':
                    autopublish = True
                else:
                    autopublish = False
                # print(autopublish)

                # ################### ----- Set Classification (single and multiple value)------ ####################
                if not sheet['AV' + str(i)].value is None:
                    classExcel = sheet['AV' + str(i)].value.replace(" ", "")
                    classExcelDict = classExcel.split("|")
                    classExcelArray = []
                    for classification in classExcelDict:
                        classificationRes = getPackageTypeApi()
                        for classs in classificationRes['classification']:
                            if classs['classificationName'] == classification:
                                classExcelArray.append("" + str(classs['classificationId']) + "")
                    pkgclassification = classExcelArray
                    # print(pkgclassification)
                else:
                    print('Classification is Empty')
                    break

                # ################### ----- Set BoxsetTypeId ------ ####################
                if not sheet['AW' + str(i)].value is None:
                    boxsetTypeRes = getPackageTypeApi()
                    for boxsetType in boxsetTypeRes['boxsetType']:
                        if boxsetType['boxsetName'] == sheet['AW' + str(i)].value:
                            boxsettypeid = boxsetType['boxsetId']
                            # print(boxsetType['boxsetId'])
                else:
                    boxsettypeid = ''

                # ################### ----- Sort Order ------ ####################
                if not sheet['AX' + str(i)].value is None:
                    sortorder = sheet['AX' + str(i)].value
                else:
                    sortorder = ''

                # ################### ----- Acquisition Start & End ------ ####################
                acqstart = ''
                acqend = ''
                if not sheet['AY' + str(i)].value is None:
                    acqstart = sheet['AY' + str(i)].value.strftime("%Y-%m-%d %H:%M:%S")

                if not sheet['AZ' + str(i)].value is None:
                    acqend = sheet['AZ' + str(i)].value.strftime("%Y-%m-%d %H:%M:%S")

                # ################### ----- Sponsored ------ ####################
                if sheet['BA' + str(i)].value == 'Y':
                    sponsored = True
                else:
                    sponsored = False

                # ################### ----- Supplier ------ ####################
                if not sheet['BB' + str(i)].value is None:
                    supplieridRes = getPackageTypeApi()
                    for supplier in supplieridRes['supplier']:
                        if supplier['supplierName'] == sheet['BB' + str(i)].value:
                            supplierid = supplier['supplierId']
                else:
                    supplierid = ''

                # ################### ----- Offer Creation ------ ####################

                # ################### ----- STB ------ ####################
                pkgoffers = []
                if not sheet['BC' + str(i)].value is None:
                    pkgOfferSTBRes = offerApi()
                    for serviceAPISTBList in pkgOfferSTBRes['service']:
                        if serviceAPISTBList['serviceLabel'] == sheet['BC' + str(i)].value:
                            serviceSTB = serviceAPISTBList['serviceId']

                    # ----- STB Offer Start------ #
                    offerStartSTB = sheet['BD' + str(i)].value.strftime("%Y-%m-%d %H:%M:%S")

                    # ----- STB Offer End------ #
                    offerEndSTB = sheet['BE' + str(i)].value.strftime("%Y-%m-%d %H:%M:%S")

                    # ----- STB PCT------ #
                    if not sheet['BF' + str(i)].value is None:
                        pctExcel = sheet['BF' + str(i)].value.replace(" ", "")
                        pctExcelDict = pctExcel.split("|")
                        pctSTBArray = []
                        for pct in pctExcelDict:
                            pctRes = offerApi()
                            for classs in pctRes['providerContentTier']:
                                if classs['providerContentTierLabel'] == pct and classs['serviceId'] == serviceSTB:
                                    pctSTBArray.append("" + str(classs['providerContentTierId']) + "")
                        pctSTB = pctSTBArray
                    else:
                        print('STB PCT is Empty')
                        break

                    # ----- STB Charge Code------ #
                    if not sheet['BG' + str(i)].value is None:
                        chgCodeSTB = sheet['BG' + str(i)].value
                    else:
                        chgCodeSTB = ''

                    # ################### ----- Business Model ------ ####################
                    bmSTB = ''
                    currencySTB = ''
                    priceSTB = ''

                    if not sheet['BH' + str(i)].value is None:
                        offerRes = offerApi()
                        offerRes2 = offerRes['offerRow']
                        for bizmodSTB in offerRes2['businessModel']:
                            if bizmodSTB['businessModelLabel'] == sheet['BH' + str(i)].value:
                                bmSTB = bizmodSTB['businessModelId']

                        # ##### --- Set currency and Price for Business Model 9 - TVOD ----- #####
                        if bmSTB == 4:
                            priceRes = offerApi()
                            priceRes2 = priceRes['offerRow']
                            for priceSTB1 in priceRes2['price']:
                                if priceSTB1['priceLabel'] == str(sheet['BI' + str(i)].value):
                                    priceSTB = priceSTB1['priceId']
                                    currencySTB = 1
                    else:
                        print('STB Business Model is Empty')

                    # ----- STB Max View------ #
                    maxViewSTB = '0'
                    if not sheet['BJ' + str(i)].value is None:
                        maxViewSTB = sheet['BJ' + str(i)].value
                    else:
                        print('STB Max View is Empty')

                # ######## ------ Package STB Offer Payload ----- #############
                pkgOfferSTB = {"serviceId": "" + str(serviceSTB) + "", "offerStart": "" + offerStartSTB + "", "offerEnd": "" + offerEndSTB + "", "providerContentTierId": pctSTB, "thirdPartyId": [], "comingSoonEndDate": "", "assetLifeCycleId": "", "sfvAccountId": [], "chargeCode": chgCodeSTB, "download2Go": False, "d2GoRetentionPeriodId": "", "d2GoPlaybackPeriodId": "", "d2GoMaxPlay_countId": "", "offerRow": [{"regionId": "1", "currencyId": currencySTB, "priceId": priceSTB, "inAppPrice": "", "bmId": bmSTB, "maxViewId": maxViewSTB}]}
                pkgoffers = [pkgOfferSTB]
                # pkgoffers = json.dumps(pkgOfferArray)
                # print(pkgoffers)

                # ################### ----- IVP ------ ####################
                if not sheet['BK' + str(i)].value is None:
                    pkgOfferIVPRes = offerApi()
                    for serviceAPIIVPList in pkgOfferIVPRes['service']:
                        if serviceAPIIVPList['serviceLabel'] == sheet['BK' + str(i)].value:
                            serviceIVP = serviceAPIIVPList['serviceId']

                    # ----- IVP Offer Start------ #
                    offerStartIVP = sheet['BL' + str(i)].value.strftime("%Y-%m-%d %H:%M:%S")

                    # ----- IVP Offer End------ #
                    offerEndIVP = sheet['BM' + str(i)].value.strftime("%Y-%m-%d %H:%M:%S")

                    # ----- IVP PCT------ #
                    if not sheet['BN' + str(i)].value is None:
                        pctExcel = sheet['BN' + str(i)].value.replace(" ", "")
                        pctExcelDict = pctExcel.split("|")
                        pctIVPArray = []
                        for pct in pctExcelDict:
                            pctRes = offerApi()
                            for classs in pctRes['providerContentTier']:
                                if classs['providerContentTierLabel'] == pct and classs['serviceId'] == serviceIVP:
                                    pctIVPArray.append("" + str(classs['providerContentTierId']) + "")
                        pctIVP = pctIVPArray
                    else:
                        print('IVP PCT is Empty')
                        break

                    # ----- IVP Charge Code------ #
                    if not sheet['BO' + str(i)].value is None:
                        chgCodeIVP = sheet['BO' + str(i)].value
                    else:
                        chgCodeIVP = ''

                    # ################### ----- Business Model ------ ####################
                    bmIVP = ''
                    currencyIVP = ''
                    priceIVP = ''

                    if not sheet['BP' + str(i)].value is None:
                        offerRes = offerApi()
                        offerRes2 = offerRes['offerRow']
                        for bizmodIVP in offerRes2['offerType']:
                            if bizmodIVP['offerTypeLabel'] == sheet['BP' + str(i)].value:
                                bmIVP = bizmodIVP['offerTypeId']

                        # ##### --- Set IVP currency and Price for Offer Type TVOD ----- #####
                        if bmIVP == 1:
                            priceRes = offerApi()
                            priceRes2 = priceRes['offerRow']
                            for priceIVP1 in priceRes2['price']:
                                if priceIVP1['priceLabel'] == str(sheet['BQ' + str(i)].value):
                                    priceIVP = priceIVP1['priceId']
                                    currencyIVP = 1

                        # ##### --- Set IVP InApp Price ----- #####
                        if not sheet['BR' + str(i)].value is None:
                            inAppPrcIVPRes = offerApi()
                            for inAppPrcIVP1 in inAppPrcIVPRes['inAppPrice']:
                                if inAppPrcIVP1['inAppPriceLabel'] == str(sheet['BR' + str(i)].value):
                                    inAppPrcIVP = inAppPrcIVP1['inAppPriceId']
                        else:
                            inAppPrcIVP = 5

                        # ----- IVP Max View------ #
                        maxViewIVP = '0'
                        if not sheet['BS' + str(i)].value is None:
                            maxViewIVP = sheet['BS' + str(i)].value
                        else:
                            maxViewIVP = '0'

                        # ----- IVP D2Go Retention Period------ #
                        if not sheet['BT' + str(i)].value is None:
                            d2gIVPRes = offerApi()
                            for d2gIVPRes1 in d2gIVPRes['d2GoRetentionPeriod']:
                                if d2gIVPRes1['downloadToGoRetentionLabel'] == sheet['BT' + str(i)].value and d2gIVPRes1['serviceId'] == serviceIVP:
                                    d2goRetentionIVP = d2gIVPRes1['downloadToGoRetentionId']
                                    d2goIVP = True

                            # ----- IVP D2Go Playback Period------ #
                            if not sheet['BU' + str(i)].value is None:
                                d2gPlaybackIVPRes = offerApi()
                                for d2gPlaybackIVPRes1 in d2gPlaybackIVPRes['d2GoPlaybackPeriod']:
                                    if d2gPlaybackIVPRes1['downloadToGoPlaybackLabel'] == sheet['BU' + str(i)].value and d2gPlaybackIVPRes1['serviceId'] == serviceIVP:
                                        d2goPlPdIVP = d2gPlaybackIVPRes1['downloadToGoPlaybackId']
                            else:
                                d2goPlPdIVP = ''

                            # ----- IVP D2Go Max Play-count------ #
                            if not sheet['BV' + str(i)].value is None:
                                d2gMaxPlaybackIVPRes = offerApi()
                                for d2gMaxPlaybackIVPRes1 in d2gMaxPlaybackIVPRes['d2GoMaxPlayCount']:
                                    if d2gMaxPlaybackIVPRes1['downloadToGoMaxPlayCountLabel'] == sheet['BV' + str(i)].value and d2gMaxPlaybackIVPRes1['serviceId'] == serviceIVP:
                                        d2goMaxPlIVP = d2gMaxPlaybackIVPRes1['downloadToGoMaxPlayCountId']
                            else:
                                d2goMaxPlIVP = ''

                        else:
                            d2goIVP = False
                            d2goRetentionIVP = ''
                            d2goPlPdIVP = ''
                            d2goMaxPlIVP = ''

                # ######## ------ Package IVP Offer Payload ----- #############
                pkgOfferIVP = {"serviceId": serviceIVP, "offerStart": offerStartIVP, "offerEnd": offerEndIVP, "providerContentTierId": pctIVP, "thirdPartyId": [], "comingSoonEndDate": "", "assetLifeCycleId": "", "sfvAccountId": [], "chargeCode": chgCodeIVP, "castingId": [], "blockAds": False, "preLogin": False, "download2Go": d2goIVP, "d2GoRetentionPeriodId": d2goRetentionIVP, "d2GoPlaybackPeriodId": d2goPlPdIVP, "d2GoMaxPlay_countId": d2goMaxPlIVP, "offerRow": [{"offerTypeId": bmIVP, "regionId": "1", "currencyId": currencyIVP, "priceId": priceIVP, "inAppPrice": inAppPrcIVP, "maxViewId": maxViewIVP}]}
                #pkgoffers = [pkgOfferIVP] pkgOfferSTB
                pkgoffers.append(pkgOfferIVP)
                #print(pkgoffers)

                # ################### ----- SOTT ------ ####################

                if not sheet['BW' + str(i)].value is None:
                    pkgOfferSOTTRes = offerApi()
                    for serviceAPISOTTList in pkgOfferSOTTRes['service']:
                        if serviceAPISOTTList['serviceLabel'] == sheet['BW' + str(i)].value:
                            serviceSOTT = serviceAPISOTTList['serviceId']

                    # ----- SOTT Offer Start------ #
                    offerStartSOTT = sheet['BX' + str(i)].value.strftime("%Y-%m-%d %H:%M:%S")

                    # ----- SOTT Offer End------ #
                    offerEndSOTT = sheet['BY' + str(i)].value.strftime("%Y-%m-%d %H:%M:%S")

                    # ----- SOTT PCT------ #
                    if not sheet['BZ' + str(i)].value is None:
                        #pctExcel = sheet['BZ' + str(i)].value.replace(" ", "")
                        pctExcel = sheet['BZ' + str(i)].value
                        pctExcelDict = pctExcel.split("|")
                        pctSOTTArray = []
                        for pct in pctExcelDict:
                            pctRes = offerApi()
                            for classs in pctRes['providerContentTier']:
                                if classs['providerContentTierLabel'] == pct and classs['serviceId'] == serviceSOTT:
                                    pctSOTTArray.append("" + str(classs['providerContentTierId']) + "")
                                    pctSOTT = pctSOTTArray
                    else:
                        pctSOTT = pctSOTTArray

                    # ----- SOTT ChargeCode------ #
                    if not sheet['CA' + str(i)].value is None:
                        chgCodeSOTT = sheet['CA' + str(i)].value
                    else:
                        chgCodeSOTT = ''

                    # ################### ----- Business Model SOTT------ ####################
                    bmSOTT = ''
                    currencySOTT = ''
                    priceSOTT = ''
                    inAppPrcSOTT = ''

                    if not sheet['CB' + str(i)].value is None:
                        offerRes = offerApi()
                        offerRes2 = offerRes['offerRow']
                        for bizmodSOTT in offerRes2['offerType']:
                            if bizmodSOTT['offerTypeLabel'] == sheet['CB' + str(i)].value:
                                bmSOTT = bizmodSOTT['offerTypeId']

                        # ##### --- Set IVP currency and Price for Offer Type TVOD ----- #####
                        if bmSOTT == 1:
                            priceRes = offerApi()
                            priceRes2 = priceRes['offerRow']
                            for priceSOTT1 in priceRes2['price']:
                                if priceSOTT1['priceLabel'] == str(sheet['CC' + str(i)].value):
                                    priceSOTT = priceSOTT1['priceId']
                                    currencySOTT = 1

                        # ##### --- Set IVP InApp Price ----- #####
                        if not sheet['CD' + str(i)].value is None:
                            inAppPrcSOTTRes = offerApi()
                            for inAppPrcSOTT1 in inAppPrcSOTTRes['inAppPrice']:
                                if inAppPrcSOTT1['inAppPriceLabel'] == str(sheet['CD' + str(i)].value):
                                    inAppPrcSOTT = inAppPrcSOTT1['inAppPriceId']
                        else:
                            inAppPrcSOTT = 5

                    # ----- IVP Max View------ #
                    maxViewSOTT = '0'
                    if not sheet['CE' + str(i)].value is None:
                        maxViewSOTT = sheet['CE' + str(i)].value
                    else:
                        maxViewSOTT = '0'

                    # ----- SOTT D2Go Retention Period------ #
                    if not sheet['CF' + str(i)].value is None:
                        d2gSOTTRes = offerApi()
                        for d2gSOTTRes1 in d2gSOTTRes['d2GoRetentionPeriod']:
                            if d2gSOTTRes1['downloadToGoRetentionLabel'] == sheet['CF' + str(i)].value and d2gSOTTRes1['serviceId'] == serviceSOTT:
                                d2goRetentionSOTT = d2gSOTTRes1['downloadToGoRetentionId']
                                d2goSOTT = True

                        # ----- IVP D2Go Playback Period------ #
                        if not sheet['CG' + str(i)].value is None:
                            d2gPlaybackSOTTRes = offerApi()
                            for d2gPlaybackSOTTRes1 in d2gPlaybackSOTTRes['d2GoPlaybackPeriod']:
                                if d2gPlaybackSOTTRes1['downloadToGoPlaybackLabel'] == sheet['CG' + str(i)].value and d2gPlaybackSOTTRes1['serviceId'] == serviceSOTT:
                                    d2goPlPdSOTT = d2gPlaybackSOTTRes1['downloadToGoPlaybackId']
                        else:
                            d2goPlPdSOTT = ''

                        # ----- IVP D2Go Max Play-count------ #
                        if not sheet['CH' + str(i)].value is None:
                            d2gMaxPlaybackSOTTRes = offerApi()
                            for d2gMaxPlaybackSOTTRes1 in d2gMaxPlaybackSOTTRes['d2GoMaxPlayCount']:
                                if d2gMaxPlaybackSOTTRes1['downloadToGoMaxPlayCountLabel'] == sheet['CH' + str(i)].value and d2gMaxPlaybackSOTTRes1['serviceId'] == serviceIVP:
                                    d2goMaxPlSOTT = d2gMaxPlaybackSOTTRes1['downloadToGoMaxPlayCountId']
                        else:
                            d2goMaxPlSOTT = ''

                    else:
                        d2goSOTT = False
                        d2goRetentionSOTT = ''
                        d2goPlPdSOTT = ''
                        d2goMaxPlSOTT = ''

                    # ----- SOTT Third Party------ #
                    if not sheet['CI' + str(i)].value is None:
                        thdPartySOTTExcel = sheet['CI' + str(i)].value.replace(" ", "")
                        thdPartySOTTExcelDict = thdPartySOTTExcel.split("|")
                        thdPartySOTTArray = []
                        for pct in thdPartySOTTExcelDict:
                            thdPartySOTTRes = offerApi()
                            for classs in thdPartySOTTRes['thirdParty']:
                                if classs['thirdPartyLabel'] == pct:
                                    thdPartySOTTArray.append("" + str(classs['thirdPartyId']) + "")
                                    thdPartySOTT = thdPartySOTTArray
                    else:
                        thdPartySOTT = []

                    # ----- SOTT comingSoonEndDate------ #
                    if not sheet['CJ' + str(i)].value is None:
                        csEndDateSOTT = sheet['CJ' + str(i)].value.strftime("%Y-%m-%d %H:%M:%S")
                    else:
                        csEndDateSOTT = ''

                    # ----- SOTT assetLifeCycle------ #
                    if not sheet['CK' + str(i)].value is None:
                        assetLCRes = offerApi()
                        for assetLCList in assetLCRes['assetLifeCycle']:
                            if assetLCList['assetLifeCycleLabel'] == sheet['CK' + str(i)].value:
                                assetLCSOTT = str(assetLCList['assetLifeCycleId'])
                    else:
                        assetLCSOTT = ''

                # ######## ------ Package SOTT Offer Payload ----- #############
                pkgOfferSOTT = {"serviceId": serviceSOTT, "offerStart": offerStartSOTT, "offerEnd": offerEndSOTT, "providerContentTierId": pctSOTT, "thirdPartyId": thdPartySOTT, "comingSoonEndDate": csEndDateSOTT, "assetLifeCycleId": assetLCSOTT, "sfvAccountId": [], "chargeCode": chgCodeSOTT, "castingId": [], "blockAds": False, "preLogin": False, "download2Go": d2goSOTT, "d2GoRetentionPeriodId": d2goRetentionSOTT, "d2GoPlaybackPeriodId": d2goPlPdSOTT, "d2GoMaxPlay_countId": d2goMaxPlSOTT, "offerRow": [{"offerTypeId": bmSOTT, "regionId": "1", "currencyId": currencySOTT, "priceId": priceSOTT, "inAppPrice": inAppPrcSOTT, "maxViewId": maxViewSOTT}]}
                pkgoffers.append(pkgOfferSOTT)
                # print(pkgOfferSOTT)

                # ############ ----- Package Metadata ----- ##########

                # ############ ----- Set Package Metadata Title ----- ##########
                pkgMtdRes = getContent(contentid)
                for pkgMtdRes1 in pkgMtdRes['contentCoreData']:
                    pkgMtdSysTit = pkgMtdRes1['systemTitle']

                # ############ ----- Set Package Metadata Certification ----- ##########
                pkgMtdCertRes = vodFilters()
                for pkgMtdCertRes1 in pkgMtdCertRes['certification']:
                    if pkgMtdCertRes1['certificationValue'] == sheet['CL' + str(i)].value:
                        pkgMtdCert = str(pkgMtdCertRes1['certificationId'])

                # ############ ----- Set Package Metadata Language ----- ##########
                pkgMtdLangCntRes = getContent(contentid)
                for pkgMtdLangCntRes1 in pkgMtdLangCntRes['contentCoreData']:
                    pkgMtdLang = str(pkgMtdLangCntRes1['originalLanguage'])

                pkgMtdLangRes = vodFilters()
                for pkgMtdLangRes1 in pkgMtdLangRes['audioLanguage']:
                    if pkgMtdLangRes1['languageLabel'] == pkgMtdLang:
                        pkgMtdLang = str(pkgMtdLangRes1['languageId'])

                # ############ ----- Set Package Metadata Year Of Release ----- ##########
                pkgMtdYORRes = getContent(contentid)
                for pkgMtdYORRes1 in pkgMtdYORRes['contentCoreData']:
                    pkgMtdYOR = str(pkgMtdYORRes1['yearOfRelease'])

                # ############ ----- Set Package Metadata Filter----- ##########
                pkgMtdFilStrRes = getContent(contentid)
                for pkgMtdFilStrRes1 in pkgMtdFilStrRes['contentCoreData']:
                    for pkgMtdFilStrRes2 in pkgMtdFilStrRes1['contentGenre']:
                        if pkgMtdFilStrRes2['genreType'] == 'Category/Filter':
                            pkgMtdStrFil = pkgMtdFilStrRes2['genre']

                pkgMtdFilRes = vodFilters()
                for pkgMtdFilRes1 in pkgMtdFilRes['filter']:
                    if pkgMtdFilRes1['filterName'] == pkgMtdStrFil:
                        pkgMtdFil = str(pkgMtdFilRes1['filterId'])

                # ############ ----- Set Package Metadata SubFilter----- ##########
                pkgMtdFil2StrArray = []
                pkgMtdFil2StrRes = getContent(contentid)
                for pkgMtdFil2StrRes1 in pkgMtdFil2StrRes['contentCoreData']:
                    for pkgMtdFil2StrRes2 in pkgMtdFil2StrRes1['contentGenre']:
                        if pkgMtdFil2StrRes2['genreType'] == 'Category/Filter':
                            pkgMtdStrFil2 = pkgMtdFil2StrRes2['subgenre']
                            # ######### --- To get Subfilter Int value ----- ########
                            pkgMtdFil2Res = vodFilters()
                            for pkgMtdFilRes1 in pkgMtdFil2Res['filter']:
                                for pkgMtdFilRes2 in pkgMtdFilRes1['subFilters']:
                                    if pkgMtdFilRes2['subFilterName'] == pkgMtdStrFil2:
                                        pkgMtdFil2StrArray.append("" + str(pkgMtdFilRes2['subFilterId']) + "")
                                        pkgMtdFil2 = pkgMtdFil2StrArray

                # ############ ----- Set Package Metadata Genre/SubGenre----- ##########
                pkgMtdGenStrRes = getContent(contentid)
                for pkgMtdGenStrRes1 in pkgMtdGenStrRes['contentCoreData']:
                    for pkgMtdGenStrRes2 in pkgMtdGenStrRes1['contentGenre']:
                        if pkgMtdGenStrRes2['genreType'] == 'DVB':
                            pkgMtdGen = str(pkgMtdGenStrRes2['genreId'])
                            pkgMtdSubGen = str(pkgMtdGenStrRes2['subGenreId'])

                # ############ ----- Set Package Metadata Short Synopsis----- ##########
                pkgMtdShtSynStrRes = getContent(contentid)
                for pkgMtdShtSynStrRes1 in pkgMtdShtSynStrRes['contentCoreData']:
                    for pkgMtdShtSynStrRes2 in pkgMtdShtSynStrRes1['contentSynopsis']:
                        if pkgMtdShtSynStrRes2['synopsisType'] == 'Short':
                            pkgMtdShtSyn = pkgMtdShtSynStrRes2['synopsis']

                        if pkgMtdShtSynStrRes2['synopsisType'] == 'Long':
                            pkgMtdLngSyn = pkgMtdShtSynStrRes2['synopsis']


                # {"englishMetadata":{"title":"Tokyo 2020 Olympic: Daily Highlights Day -2 Part 1","certification":"3","audioLanguage":"5","yearOfRealease":"2021","filter":"8","subFilter":["57","58"],"genre":"4","subGenre":"11","shortSynopsis":"Catch up on Olympics Tokyo 2020 daily Highlights, actions and best moments - Day -02 Part 1","longSynopsis":"Catch up on Olympics Tokyo 2020 daily Highlights, actions and best moments - Day -02 Part 1","contentId":"12931"},"vernacular":[{"languageId":"1"},{"languageId":"2"},{"languageId":"3"},{"languageId":"4"},{"languageId":"5"}]}
                metadata = {"englishMetadata": {"title": pkgMtdSysTit, "certification": pkgMtdCert, "audioLanguage": pkgMtdLang, "yearOfRealease": pkgMtdYOR, "filter": pkgMtdFil, "subFilter": pkgMtdFil2, "genre": pkgMtdGen, "subGenre": pkgMtdSubGen, "shortSynopsis": pkgMtdShtSyn, "longSynopsis": pkgMtdShtSyn, "contentId": contentid}, "vernacular": [{"languageId": "1"}, {"languageId": "2"}, {"languageId": "3"}, {"languageId": "4"}, {"languageId": "5"}]}
                print(metadata)

                # ########## savePackageapi API call function ###############
                packageRes = savePackageapi(packagename, pkgid, pkgtypeid, autopublish, channelowner, pkgclassification, boxsettypeid, sortorder, acqstart, acqend, sponsored, supplierid, contentid, pkgoffers, metadata)
                print(packageRes['packId'])
                #print(packageRes)

            # #################--- (MP) Season ---##################
            elif not sheet['J' + str(i)].value is None and sheet['K' + str(i)].value is None:
                print('Season')

            # #################--- (MP) Episode ---##################
            else:
                print('Episode')

        # #################--- IF status = M ---##################
        elif sheet['A' + str(i)].value == 'M':

            # #################--- (M) ACID Empty ---##################
            if sheet['C' + str(i)].value is None:
                print('M ACID Empty')

            # #################--- (M) ACID Not-Empty ---##################
            else:
                print('M ACID not Empty')

        # #################--- IF status = P ---##################
        elif sheet['A' + str(i)].value == 'P':

            # #################--- (P) ACID Empty ---##################
            if sheet['C' + str(i)].value is None:
                print('P ACID Empty')

            # #################--- (P) ACID Not-Empty ---##################
            else:
                print('P ACID not Empty')

        # #################--- Print error if wrong value ---##################
        else:
            print('Wrong value for status')

    wb.save('AMS_Status.xlsx')
